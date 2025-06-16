
import pandas as pd
import io
from typing import Union, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class SpreadsheetProcessor:
    """Handles loading and processing of spreadsheet files"""
    
    def __init__(self):
        """Initialize the spreadsheet processor"""
        self.supported_formats = ['.csv', '.xlsx', '.xls']
    
    def load_file(self, uploaded_file) -> pd.DataFrame:
        """
        Load a spreadsheet file into a pandas DataFrame
        
        Args:
            uploaded_file: Flask FileStorage object or Streamlit uploaded file object
            
        Returns:
            pandas DataFrame containing the spreadsheet data
        """
        if uploaded_file is None:
            raise ValueError("No file provided")
        
        # Handle both Flask FileStorage and Streamlit UploadedFile
        filename = getattr(uploaded_file, 'filename', '') or getattr(uploaded_file, 'name', '')
        
        if not filename:
            raise ValueError("Could not determine filename")
        
        file_extension = self.get_file_extension(filename)
        
        if file_extension not in self.supported_formats:
            raise ValueError(f"Unsupported file format '{file_extension}'. Supported formats: {', '.join(self.supported_formats)}")
        
        try:
            # Reset file pointer to beginning
            if hasattr(uploaded_file, 'seek'):
                uploaded_file.seek(0)
            
            # Read file content into BytesIO for consistent handling
            if hasattr(uploaded_file, 'read'):
                file_content = uploaded_file.read()
            else:
                file_content = uploaded_file.getvalue()
            
            file_buffer = io.BytesIO(file_content)
            
            if file_extension == '.csv':
                # Try different encodings for CSV files
                encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                
                for encoding in encodings:
                    try:
                        file_buffer.seek(0)
                        df = pd.read_csv(file_buffer, encoding=encoding)
                        return self.clean_dataframe(df)
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                
                # If all encodings fail, try with error handling
                file_buffer.seek(0)
                df = pd.read_csv(file_buffer, encoding='utf-8', errors='replace')
                return self.clean_dataframe(df)
                
            elif file_extension in ['.xlsx', '.xls']:
                # Handle Excel files
                try:
                    file_buffer.seek(0)
                    engine = 'openpyxl' if file_extension == '.xlsx' else 'xlrd'
                    df = pd.read_excel(file_buffer, engine=engine)
                    return self.clean_dataframe(df)
                except Exception as e:
                    # Try with different engine if the first one fails
                    try:
                        file_buffer.seek(0)
                        df = pd.read_excel(file_buffer)  # Let pandas choose engine
                        return self.clean_dataframe(df)
                    except Exception as e2:
                        raise Exception(f"Failed to read Excel file: {str(e)} / {str(e2)}")
            
        except Exception as e:
            raise Exception(f"Error loading file: {str(e)}")
    
    def clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and normalize the DataFrame
        
        Args:
            df: Input DataFrame
            
        Returns:
            Cleaned DataFrame
        """
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Clean column names - remove extra whitespace and make consistent
        df.columns = df.columns.astype(str).str.strip()
        
        # Reset index
        df = df.reset_index(drop=True)
        
        return df
    
    def get_file_extension(self, filename: str) -> str:
        """
        Get file extension from filename
        
        Args:
            filename: Name of the file
            
        Returns:
            File extension (including the dot)
        """
        if not filename or '.' not in filename:
            return ''
        
        # Handle cases where filename might have multiple dots
        extension = '.' + filename.split('.')[-1].lower()
        return extension
    
    def validate_columns(self, df: pd.DataFrame, required_columns: list) -> dict:
        """
        Validate that required columns exist in the DataFrame
        
        Args:
            df: DataFrame to validate
            required_columns: List of required column names
            
        Returns:
            Dictionary with validation results
        """
        available_columns = df.columns.tolist()
        missing_columns = []
        
        for col in required_columns:
            if col not in available_columns:
                missing_columns.append(col)
        
        return {
            'valid': len(missing_columns) == 0,
            'missing_columns': missing_columns,
            'available_columns': available_columns
        }
    
    def suggest_url_column(self, df: pd.DataFrame) -> Optional[str]:
        """
        Suggest which column might contain URLs
        
        Args:
            df: DataFrame to analyze
            
        Returns:
            Suggested column name or None
        """
        # Common URL column names
        url_column_patterns = [
            'url', 'link', 'website', 'web', 'href', 'uri', 'address',
            'shortened_url', 'short_url', 'shorturl', 'link_url'
        ]
        
        # Check for exact matches first
        for col in df.columns:
            if col.lower().strip() in url_column_patterns:
                return col
        
        # Check for partial matches
        for col in df.columns:
            col_lower = col.lower().strip()
            for pattern in url_column_patterns:
                if pattern in col_lower or col_lower in pattern:
                    return col
        
        # Check column content for URL-like patterns
        for col in df.columns:
            if df[col].dtype == 'object':  # String columns only
                sample_values = df[col].dropna().head(10).astype(str)
                url_count = 0
                
                for value in sample_values:
                    value_lower = value.lower().strip()
                    if any(pattern in value_lower for pattern in ['http', 'www.', '.com', '.org', '.net', 'bit.ly', 'tinyurl']):
                        url_count += 1
                
                # If more than 50% of sample values look like URLs
                if len(sample_values) > 0 and (url_count / len(sample_values)) > 0.5:
                    return col
        
        return None
    
    def export_to_excel(self, df: pd.DataFrame, filename: str = None) -> bytes:
        """
        Export DataFrame to Excel format with formatting
        
        Args:
            df: DataFrame to export
            filename: Optional filename
            
        Returns:
            Excel file as bytes
        """
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Processed URLs')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Processed URLs']
            
            # Apply formatting
            self.format_excel_worksheet(worksheet, df)
        
        output.seek(0)
        return output.getvalue()
    
    def format_excel_worksheet(self, worksheet, df: pd.DataFrame):
        """
        Apply formatting to Excel worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame being exported
        """
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format header row
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding, but not too wide
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add conditional formatting for status column
        if 'status' in df.columns:
            status_col_index = df.columns.get_loc('status') + 1
            
            # Success cells - green background
            success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for row_num in range(2, len(df) + 2):  # Skip header row
                cell = worksheet.cell(row=row_num, column=status_col_index)
                if cell.value == 'Success':
                    cell.fill = success_fill
                elif cell.value in ['Failed', 'Error']:
                    cell.fill = error_fill
    
    def get_dataframe_info(self, df: pd.DataFrame) -> dict:
        """
        Get comprehensive information about the DataFrame
        
        Args:
            df: DataFrame to analyze
            
        Returns:
            Dictionary with DataFrame information
        """
        info = {
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'column_names': df.columns.tolist(),
            'memory_usage': df.memory_usage(deep=True).sum(),
            'data_types': df.dtypes.to_dict(),
            'missing_values': df.isnull().sum().to_dict(),
            'suggested_url_column': self.suggest_url_column(df)
        }
        
        # Add column statistics for numeric columns
        numeric_columns = df.select_dtypes(include=['number']).columns
        if len(numeric_columns) > 0:
            info['numeric_summary'] = df[numeric_columns].describe().to_dict()
        
        return info
